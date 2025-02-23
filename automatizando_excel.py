import openpyxl
import openpyxl.workbook

# criar uma Pagina
book = openpyxl.Workbook()

# Visualiza pagina existente
print (book.sheetnames)

# Criar uma Planilha
book.create_sheet('Computadores')

# selecionar uma pagina 
Computadores_page = book ['Computadores']
Computadores_page.append(['Eletronica','Memória Ram','Preço'])
Computadores_page.append(['Computador1 ','8GB','R$ 2.000'])
Computadores_page.append(['Computador2','20GB','R$ 4.000'])
Computadores_page.append(['Computador3','16GB','R$ 1.700'])
Computadores_page.append(['Computador4','32GB','R$ 2.000'])

# salvar a planilha
book.save('Meus Computadores.xlsx')